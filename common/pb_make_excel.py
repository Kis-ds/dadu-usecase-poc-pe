from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, colors, Color, Font, Border, Side
import pandas as pd
from datetime import date

wb = Workbook()
ws1 = wb.active
ws1.title = '1. WICS별 변화율'
ws2 = wb.create_sheet('2. 특정 공시')
ws3 = wb.create_sheet('3. 특정 키워드 뉴스')
ws4 = wb.create_sheet('1-2. 종목별 변화율')


def pb_make_excel():

    task1_pd = pd.read_pickle(r'C:\Users\Administrator\PycharmProjects\datasolution_news_3_8\pbproject\task1\krx_change_last.pkl')
    ws = ws1
    ws.insert_rows(0)
    ws.freeze_panes = 'A3'

    for r in dataframe_to_rows(task1_pd, index=False, header=True):
        ws.append(r)

    for column_cells in ws.columns:
        for cell in ws[column_cells[0].column_letter]:
            cell.font = Font(size=9)
            if column_cells[0].column_letter in ['A', 'B', 'C', 'D', 'E'] :
                cell.alignment = Alignment(horizontal='center')
    for cell in ws["1"]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, bold=True)
        ws['I1'] = '산업별 평균 주가 대비 개별 종목 변화율 차이(%p)'
        green = PatternFill(start_color = 'd8e4bc', end_color = 'd8e4bc', patternType='solid')
        cell.fill=green
    for cell in ws["2"]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, bold=True)
        cell.border = Border(bottom = Side(border_style="thin"))
        green = PatternFill(start_color = 'd8e4bc', end_color = 'd8e4bc', patternType='solid')
        cell.fill=green
    ws.column_dimensions['A'].width = ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = ws.column_dimensions['G'].width = ws.column_dimensions['H'].width = 7
    ws.column_dimensions['I'].width = ws.column_dimensions['J'].width = ws.column_dimensions['K'].width = ws.column_dimensions['L'].width = 7

    task2_pd = pd.read_pickle(r'C:\Users\Administrator\PycharmProjects\datasolution_news_3_8\pbproject\task2\task2.pkl')
    ws = ws2
    ws.freeze_panes = 'A2'
    for r in dataframe_to_rows(task2_pd, index=False, header=True):
        ws.append(r)
    for column_cells in ws.columns:
        for cell in ws[column_cells[0].column_letter]:
            cell.font = Font(size=9)
            if column_cells[0].column_letter == 'F':
                cell.value = '=HYPERLINK("{}", "{}")'.format(cell.value, '링크') if cell.value != '' else ''
                cell.font = Font(size=9, italic=True, underline='singleAccounting', color='0000ff')
            if column_cells[0].column_letter in ['A', 'B', 'C', 'D', 'F']:
                cell.alignment = Alignment(horizontal='center')
    for cell in ws["1"]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, bold=True)
        cell.border = Border(bottom=Side(border_style="thin"))
        green = PatternFill(start_color='d8e4bc', end_color='d8e4bc', patternType='solid')
        cell.fill = green
    ws.column_dimensions['A'].width = ws.column_dimensions['G'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['E'].width = 85
    ws.column_dimensions['F'].width = 7

    task3_pd = pd.read_pickle(r'C:\Users\Administrator\PycharmProjects\datasolution_news_3_8\pbproject\task3\task3.pkl')
    task3_pd = task3_pd[['종목코드', '종목명', '언론사URL', '작성일자', '제목', '네이버링크', '카테고리']]
    ws = ws3
    ws.freeze_panes = 'A2'

    for r in dataframe_to_rows(task3_pd, index=False, header=True):
        ws.append(r)
    for column_cells in ws.columns:
        for cell in ws[column_cells[0].column_letter]:
            cell.font = Font(size=9)
            if column_cells[0].column_letter == 'F':
                cell.value = '=HYPERLINK("{}", "{}")'.format(cell.value,'링크') if cell.value != '' else ''
                cell.font = Font(size=9, italic=True, underline='singleAccounting', color='0000ff')
            if column_cells[0].column_letter in ['A', 'B', 'C', 'D', 'F', 'G']:
                cell.alignment = Alignment(horizontal='center')
    for cell in ws["1"]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, bold=True)
        cell.border = Border(bottom=Side(border_style="thin"))
        green = PatternFill(start_color='d8e4bc', end_color='d8e4bc', patternType='solid')
        cell.fill = green
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 80
    ws.column_dimensions['F'].width = 7
    ws.column_dimensions['G'].width = 20

    task4_pd = pd.read_pickle(
        r'C:\Users\Administrator\PycharmProjects\datasolution_news_3_8\pbproject\task1\krx_each_stock_change.pkl')


    ws = ws4
    ws.insert_rows(0)
    ws.freeze_panes = 'A3'

    for r in dataframe_to_rows(task4_pd, index=False, header=True):
        ws.append(r)

    for column_cells in ws.columns:
        for cell in ws[column_cells[0].column_letter]:
            cell.font = Font(size=9)
            if column_cells[0].column_letter in ['A', 'B', 'C', 'D', 'E'] :
                cell.alignment = Alignment(horizontal='center')
    for cell in ws["1"]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, bold=True)
        ws['I1'] = '개별 종목 변화율(%)'
        ws['P1'] = '산업 변화율(%)'
        green = PatternFill(start_color = 'd8e4bc', end_color = 'd8e4bc', patternType='solid')
        cell.fill=green
    for cell in ws["2"]:
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, bold=True)
        cell.border = Border(bottom = Side(border_style="thin"))
        green = PatternFill(start_color = 'd8e4bc', end_color = 'd8e4bc', patternType='solid')
        cell.fill=green

        ws['M2'] = '1D'
        ws['N2'] = '1W'
        ws['O2'] = '1M'
        ws['P2'] = '2M'
        ws['Q2'] = '3M'
        ws['R2'] = '6M'
        ws['S2'] = '1Y'

    ws.column_dimensions['A'].width = ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = ws.column_dimensions['G'].width = ws.column_dimensions['H'].width = 7
    ws.column_dimensions['I'].width = ws.column_dimensions['J'].width = ws.column_dimensions['K'].width = ws.column_dimensions['L'].width = 7
    ws.column_dimensions['M'].width = ws.column_dimensions['N'].width = ws.column_dimensions['O'].width = 7
    ws.column_dimensions['P'].width = ws.column_dimensions['Q'].width = ws.column_dimensions['R'].width = ws.column_dimensions['S'].width = 7

    current_dt = date.today().strftime('%Y-%m-%d')
    wb.save(fr'C:\Users\Administrator\PycharmProjects\datasolution_news_3_8\pbproject\final_output\PB전략부_데일리모니터링_{current_dt}.xlsx')
    print("made excel file")

#pb_make_excel()

